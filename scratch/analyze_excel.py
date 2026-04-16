import openpyxl
import os

file_path = './docs/Efectividad del onboarding actualizada (version 2).xlsx'

if not os.path.exists(file_path):
    print(f"Error: File not found at {file_path}")
    exit(1)

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
except Exception as e:
    print(f"Error loading workbook: {e}")
    exit(1)

sheet_names = wb.sheetnames
print(f"Sheet names: {sheet_names}")

# Analyze first 4 sheets
for name in sheet_names[:4]:
    print(f"\n{'='*20} SHEET: {name} {'='*20}")
    sheet = wb[name]
    
    # Get first 10 rows and 20 columns to understand structure and formulas
    for r in range(1, 15):
        row_data = []
        for c in range(1, 25):
            cell = sheet.cell(row=r, column=c)
            val = cell.value
            # If it's a formula, show the formula
            if isinstance(val, str) and val.startswith('='):
                row_data.append(f"F:{val}")
            else:
                row_data.append(val)
        print(row_data)

print("\n--- Summary of interesting columns (searching for IGEO or Scores) ---")
# Search for specific keywords in headers
for name in sheet_names[:4]:
    sheet = wb[name]
    for c in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=c).value
        if header and any(kw in str(header).upper() for kw in ['IGEO', 'SCORE', 'PROMEDIO', 'TOTAL', '%']):
            print(f"[{name}] Column {c}: {header} (Sample Row 2: {sheet.cell(row=2, column=c).value})")
